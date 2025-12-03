from flask import render_template, redirect, url_for, request, flash, current_app, send_file, Response
from flask_login import login_required, current_user
from . import schedules_bp
from ..extensions import db
from ..models import Schedule, ScheduleItem, SiteSettings
from ..forms.schedules import ScheduleForm, ScheduleItemForm
from ..forms.settings import SettingsForm
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, time
from io import BytesIO


@schedules_bp.route("/")
@login_required
def list_schedules():
    schedules = Schedule.query.order_by(Schedule.date.desc()).all()
    return render_template("schedules/index.html", schedules=schedules)


@schedules_bp.route("/new", methods=["GET", "POST"])
@login_required
def new_schedule():
    form = ScheduleForm()
    if form.validate_on_submit():
        if form.is_active.data:
            Schedule.query.update({Schedule.is_active: False})
        sched = Schedule(
            name=form.name.data, 
            date=form.date.data, 
            is_active=form.is_active.data,
            show_name=form.show_name.data if form.show_name.data is not None else True,
            created_by=current_user.id
        )
        db.session.add(sched)
        db.session.commit()
        flash("Schedule created", "success")
        return redirect(url_for("schedules.edit_schedule", schedule_id=sched.id))
    return render_template("schedules/schedule_form.html", form=form, title="New Schedule")


@schedules_bp.route("/<int:schedule_id>/edit", methods=["GET", "POST"])
@login_required
def edit_schedule(schedule_id: int):
    sched = Schedule.query.get_or_404(schedule_id)
    form = ScheduleForm(obj=sched)
    if form.validate_on_submit():
        if form.is_active.data and not sched.is_active:
            Schedule.query.update({Schedule.is_active: False})
        form.populate_obj(sched)
        db.session.commit()
        flash("Schedule updated", "success")
        return redirect(url_for("schedules.list_schedules"))
    items = ScheduleItem.query.filter_by(schedule_id=schedule_id).order_by(ScheduleItem.start_time).all()
    return render_template("schedules/schedule_form.html", form=form, title="Edit Schedule", schedule=sched, items=items)


@schedules_bp.route("/<int:schedule_id>/delete", methods=["POST"])
@login_required
def delete_schedule(schedule_id: int):
    sched = Schedule.query.get_or_404(schedule_id)
    db.session.delete(sched)
    db.session.commit()
    flash("Schedule deleted", "success")
    return redirect(url_for("schedules.list_schedules"))


@schedules_bp.route("/<int:schedule_id>/items/new", methods=["GET", "POST"])
@login_required
def new_item(schedule_id: int):
    sched = Schedule.query.get_or_404(schedule_id)
    form = ScheduleItemForm()
    if form.validate_on_submit():
        item = ScheduleItem(
            schedule_id=schedule_id,
            start_time=form.start_time.data,
            duration_minutes=form.duration_minutes.data,
            location=form.location.data or None,
            uniform=form.uniform.data or None,
            lead=form.lead.data or None,
            notes=form.notes.data or None,
            icon=form.icon.data or None,
        )
        item.compute_end_time()
        db.session.add(item)
        db.session.commit()
        flash("Item added", "success")
        return redirect(url_for("schedules.edit_schedule", schedule_id=schedule_id))
    return render_template("schedules/item_form.html", form=form, schedule=sched, title="New Item")


@schedules_bp.route("/items/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
def edit_item(item_id: int):
    item = ScheduleItem.query.get_or_404(item_id)
    form = ScheduleItemForm(obj=item)
    if form.validate_on_submit():
        form.populate_obj(item)
        item.compute_end_time()
        db.session.commit()
        flash("Item updated", "success")
        return redirect(url_for("schedules.edit_schedule", schedule_id=item.schedule_id))
    return render_template("schedules/item_form.html", form=form, schedule=item.schedule, title="Edit Item")


@schedules_bp.route("/items/<int:item_id>/delete", methods=["POST"])
@login_required
def delete_item(item_id: int):
    item = ScheduleItem.query.get_or_404(item_id)
    schedule_id = item.schedule_id
    db.session.delete(item)
    db.session.commit()
    flash("Item deleted", "success")
    return redirect(url_for("schedules.edit_schedule", schedule_id=schedule_id))


@schedules_bp.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    settings = SiteSettings.query.first()
    if not settings:
        settings = SiteSettings()
        db.session.add(settings)
        db.session.commit()
    form = SettingsForm(obj=settings)
    if form.validate_on_submit():
        # handle logo upload
        file = form.logo.data
        if file:
            os.makedirs(current_app.config["UPLOAD_FOLDER"], exist_ok=True)
            filename = secure_filename(file.filename)
            path = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
            file.save(path)
            # store relative path under app/static for serving
            rel = os.path.relpath(path, os.path.join(current_app.root_path, "static"))
            settings.logo_path = rel.replace("\\", "/")
        if form.logo_size.data:
            settings.logo_size = form.logo_size.data
        # handle background image upload
        bg_file = form.background_image.data
        if bg_file:
            os.makedirs(current_app.config["UPLOAD_FOLDER"], exist_ok=True)
            filename = secure_filename(bg_file.filename)
            path = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
            bg_file.save(path)
            # store relative path under app/static for serving
            rel = os.path.relpath(path, os.path.join(current_app.root_path, "static"))
            settings.background_image_path = rel.replace("\\", "/")
        if form.background_image_size.data:
            settings.background_image_size = form.background_image_size.data
        settings.notes_left_col = form.notes_left_col.data
        settings.latitude = form.latitude.data
        settings.longitude = form.longitude.data
        settings.timezone = form.timezone.data or current_app.config.get("TIMEZONE")
        settings.bg_color = form.bg_color.data or "#000000"
        settings.text_color = form.text_color.data or "#ffffff"
        settings.box_color = form.box_color.data or "#212529"
        settings.box_opacity = form.box_opacity.data if form.box_opacity.data is not None else 1.0
        settings.schedule_color = form.schedule_color.data or "#212529"
        settings.schedule_opacity = form.schedule_opacity.data if form.schedule_opacity.data is not None else 1.0
        db.session.commit()
        flash("Settings saved", "success")
        return redirect(url_for("schedules.settings"))
    return render_template("schedules/settings.html", form=form, settings=settings)


@schedules_bp.route("/<int:schedule_id>/export", methods=["GET"])
@login_required
def export_schedule(schedule_id: int):
    schedule = Schedule.query.get_or_404(schedule_id)
    items = ScheduleItem.query.filter_by(schedule_id=schedule_id).order_by(ScheduleItem.start_time).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Header row
    headers = ["Name", "Start Time", "End Time", "Duration (min)", "Location", "Uniform", "Lead", "Notes", "Icon"]
    ws.append(["Schedule Name:", schedule.name])
    ws.append(["Date:", schedule.date.strftime("%Y-%m-%d")])
    ws.append(["Active:", "Yes" if schedule.is_active else "No"])
    ws.append(["Show Name:", "Yes" if schedule.show_name else "No"])
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[6]:  # Row 6 is the header row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for item in items:
        row = [
            item.name or "",
            item.start_time.strftime("%H:%M") if item.start_time else "",
            item.end_time.strftime("%H:%M") if item.end_time else "",
            item.duration_minutes or "",
            item.location or "",
            item.uniform or "",
            item.lead or "",
            item.notes or "",
            item.icon or ""
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"schedule_{schedule.name.replace(' ', '_')}_{schedule.date.strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@schedules_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_schedule():
    if request.method == "POST":
        if "file" not in request.files:
            flash("No file selected", "error")
            return redirect(url_for("schedules.import_schedule"))
        
        file = request.files["file"]
        if file.filename == "":
            flash("No file selected", "error")
            return redirect(url_for("schedules.import_schedule"))
        
        if not file.filename.endswith((".xlsx", ".xls")):
            flash("Invalid file type. Please upload an Excel file (.xlsx or .xls)", "error")
            return redirect(url_for("schedules.import_schedule"))
        
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # Read schedule metadata (first few rows)
            schedule_name = ws.cell(row=1, column=2).value or "Imported Schedule"
            date_str = ws.cell(row=2, column=2).value
            is_active = str(ws.cell(row=3, column=2).value).lower() == "yes"
            show_name = str(ws.cell(row=4, column=2).value).lower() == "yes"
            
            # Parse date
            if isinstance(date_str, datetime):
                schedule_date = date_str.date()
            elif isinstance(date_str, str):
                try:
                    schedule_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                except:
                    schedule_date = datetime.now().date()
            else:
                schedule_date = datetime.now().date()
            
            # Create schedule
            if is_active:
                Schedule.query.update({Schedule.is_active: False})
            
            schedule = Schedule(
                name=schedule_name,
                date=schedule_date,
                is_active=is_active,
                show_name=show_name,
                created_by=current_user.id
            )
            db.session.add(schedule)
            db.session.flush()  # Get the ID
            
            # Read items (starting from row 7, after headers)
            headers = [cell.value for cell in ws[6]]
            name_idx = headers.index("Name") if "Name" in headers else None
            start_idx = headers.index("Start Time") if "Start Time" in headers else None
            end_idx = headers.index("End Time") if "End Time" in headers else None
            duration_idx = headers.index("Duration (min)") if "Duration (min)" in headers else None
            location_idx = headers.index("Location") if "Location" in headers else None
            uniform_idx = headers.index("Uniform") if "Uniform" in headers else None
            lead_idx = headers.index("Lead") if "Lead" in headers else None
            notes_idx = headers.index("Notes") if "Notes" in headers else None
            icon_idx = headers.index("Icon") if "Icon" in headers else None
            
            items_imported = 0
            for row in ws.iter_rows(min_row=7, values_only=False):
                # Skip empty rows
                if not any(cell.value for cell in row):
                    continue
                
                # Parse start time
                start_time_val = row[start_idx].value if start_idx is not None else None
                if start_time_val:
                    if isinstance(start_time_val, time):
                        start_time = start_time_val
                    elif isinstance(start_time_val, datetime):
                        start_time = start_time_val.time()
                    elif isinstance(start_time_val, str):
                        try:
                            start_time = datetime.strptime(start_time_val, "%H:%M").time()
                        except:
                            continue
                    else:
                        continue
                else:
                    continue
                
                # Parse end time
                end_time = None
                if end_idx is not None and row[end_idx].value:
                    end_time_val = row[end_idx].value
                    if isinstance(end_time_val, time):
                        end_time = end_time_val
                    elif isinstance(end_time_val, datetime):
                        end_time = end_time_val.time()
                    elif isinstance(end_time_val, str):
                        try:
                            end_time = datetime.strptime(end_time_val, "%H:%M").time()
                        except:
                            pass
                
                # Parse duration
                duration = None
                if duration_idx is not None and row[duration_idx].value:
                    try:
                        duration = int(row[duration_idx].value)
                    except:
                        pass
                
                item = ScheduleItem(
                    schedule_id=schedule.id,
                    name=row[name_idx].value if name_idx is not None and row[name_idx].value else None,
                    start_time=start_time,
                    end_time=end_time,
                    duration_minutes=duration,
                    location=row[location_idx].value if location_idx is not None and row[location_idx].value else None,
                    uniform=row[uniform_idx].value if uniform_idx is not None and row[uniform_idx].value else None,
                    lead=row[lead_idx].value if lead_idx is not None and row[lead_idx].value else None,
                    notes=row[notes_idx].value if notes_idx is not None and row[notes_idx].value else None,
                    icon=row[icon_idx].value if icon_idx is not None and row[icon_idx].value else None,
                )
                
                if not item.end_time and item.duration_minutes:
                    item.compute_end_time()
                
                db.session.add(item)
                items_imported += 1
            
            db.session.commit()
            flash(f"Schedule imported successfully. {items_imported} items added.", "success")
            return redirect(url_for("schedules.edit_schedule", schedule_id=schedule.id))
        
        except Exception as e:
            db.session.rollback()
            flash(f"Error importing schedule: {str(e)}", "error")
            return redirect(url_for("schedules.import_schedule"))
    
    return render_template("schedules/import.html")


